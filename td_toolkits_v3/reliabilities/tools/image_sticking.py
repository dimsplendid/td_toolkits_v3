from __future__ import annotations

import re
from pathlib import Path
from enum import Enum

from typing import Dict, List,  Optional
from pydantic import BaseModel, Field
from datetime import timedelta

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

class ISType(Enum):
    S = 'Surface'
    L = 'Line'

class Color(Enum):
    R = 'D82526'
    Y = 'FFC156'
    G = '69B764'
    W = 'FFFFFF'
    K = '000000'
    WY = 'FFDD71'
    
class CellPointer(BaseModel):
    row: int = Field(1, ge=1)
    col: int = Field(1, ge=1)
    
    def __iter__(self):
        return iter((self.row, self.col))

class CheckPoint(BaseModel):
    stress_time: timedelta
    recover_time: timedelta
    gray_level: int = Field(..., ge=0, le=255)
    is_type: ISType
        
class ImageStickingBase(CheckPoint):
    ra_level: int
    
class Spec(ImageStickingBase):
    remark: str = ''
    
class Judgement(Spec):
    ok: int = Field(default=0, ge=0)
    ng: int = Field(default=0, ge=0)
    ng_level: int = Field(default=0, ge=0, le=6)
    
    @property
    def total(self):
        return self.ok + self.ng
    
    @property
    def yeild(self):
        return f"{self.ok}/{self.total}"
    
    @property
    def is_pass(self):
        if self.total == 0:
            return None
        elif self.ng == 0:
            return True
        else:
            return False
        
class Chip(BaseModel):
    name: str
    remark: Optional[str] = None
    condition: str
    
class Log(Chip):
    value: List[ImageStickingBase] = []
        
class Judger(BaseModel):
    name: str
    judgements: List[Judgement] = []
    
    def judge(self, logs: List[Log]):
        for judgement in self.judgements:
            for log in logs:
                if log.value is None:
                    continue
                for item in filter(
                    lambda x: (
                        x.gray_level == judgement.gray_level and
                        x.stress_time == judgement.stress_time and
                        x.is_type == judgement.is_type and
                        x.recover_time == judgement.recover_time
                    ),
                    log.value
                ):
                    if judgement.ra_level >= item.ra_level:
                        judgement.ok += 1
                    else:
                        judgement.ng += 1
                    if judgement.ng_level < item.ra_level:
                        judgement.ng_level = item.ra_level
    
    @classmethod
    def load_specs(
        cls, 
        path: Optional[Path] = None
    ) -> Dict[str, Judger]:
        dummy_judgers = {
            'INX': Judger(
                name='INX',
                judgements=[
                    Judgement(
                        stress_time=timedelta(seconds=10),
                        recover_time=timedelta(0),
                        gray_level=0,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(seconds=10),
                        recover_time=timedelta(0),
                        gray_level=0,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(seconds=10),
                        recover_time=timedelta(0),
                        gray_level=39,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(seconds=10),
                        recover_time=timedelta(0),
                        gray_level=39,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(seconds=10),
                        recover_time=timedelta(0),
                        gray_level=64,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(seconds=10),
                        recover_time=timedelta(0),
                        gray_level=64,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(seconds=10),
                        recover_time=timedelta(0),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(seconds=10),
                        recover_time=timedelta(0),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=3),
                        recover_time=timedelta(0),
                        gray_level=0,
                        is_type=ISType.S,
                        ra_level=2,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=3),
                        recover_time=timedelta(0),
                        gray_level=0,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=3),
                        recover_time=timedelta(0),
                        gray_level=39,
                        is_type=ISType.S,
                        ra_level=2,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=3),
                        recover_time=timedelta(0),
                        gray_level=39,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=3),
                        recover_time=timedelta(0),
                        gray_level=64,
                        is_type=ISType.S,
                        ra_level=2,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=3),
                        recover_time=timedelta(0),
                        gray_level=64,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=3),
                        recover_time=timedelta(0),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=2,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=3),
                        recover_time=timedelta(0),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(0),
                        gray_level=0,
                        is_type=ISType.S,
                        ra_level=1,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(0),
                        gray_level=0,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(0),
                        gray_level=39,
                        is_type=ISType.S,
                        ra_level=1,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(0),
                        gray_level=39,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(0),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=1,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(0),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=10),
                        recover_time=timedelta(seconds=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=10),
                        recover_time=timedelta(seconds=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=1),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=1),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=4),
                        recover_time=timedelta(0),
                        gray_level=0,
                        is_type=ISType.S,
                        ra_level=1,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=4),
                        recover_time=timedelta(0),
                        gray_level=0,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=4),
                        recover_time=timedelta(0),
                        gray_level=39,
                        is_type=ISType.S,
                        ra_level=1,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=4),
                        recover_time=timedelta(0),
                        gray_level=39,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=4),
                        recover_time=timedelta(0),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=1,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=4),
                        recover_time=timedelta(0),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=6),
                        recover_time=timedelta(seconds=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=3,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=6),
                        recover_time=timedelta(seconds=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=24),
                        recover_time=timedelta(minutes=5),
                        gray_level=0,
                        is_type=ISType.S,
                        ra_level=2,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=24),
                        recover_time=timedelta(minutes=5),
                        gray_level=0,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=24),
                        recover_time=timedelta(minutes=5),
                        gray_level=39,
                        is_type=ISType.S,
                        ra_level=2,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=24),
                        recover_time=timedelta(minutes=5),
                        gray_level=39,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=24),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=2,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=24),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'MP: Huawei': Judger(
                name='MP Huawei',
                judgements=[
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'MP: VIVO': Judger(
                name='MP VIVO',
                judgements=[
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(seconds=20),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(seconds=20),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'MP: 傳音': Judger(
                name='MP: 傳音',
                judgements=[
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(seconds=2),
                        gray_level=32,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(seconds=2),
                        gray_level=32,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'MP: OPPO': Judger(
                name='MP: OPPO',
                judgements=[
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(seconds=10),
                        gray_level=32,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(seconds=10),
                        gray_level=32,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(seconds=10),
                        gray_level=64,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(seconds=10),
                        gray_level=64,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'MP: 聞泰': Judger(
                name='MP: 聞泰',
                judgements=[
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(seconds=10),
                        gray_level=32,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(seconds=10),
                        gray_level=32,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(seconds=10),
                        gray_level=64,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(seconds=10),
                        gray_level=64,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'MP: 華勤': Judger(
                name='MP: 華勤',
                judgements=[
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(minutes=3),
                        gray_level=32,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(minutes=3),
                        gray_level=32,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(minutes=3),
                        gray_level=64,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(minutes=3),
                        gray_level=64,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=3),
                        gray_level=32,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=3),
                        gray_level=32,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=3),
                        gray_level=64,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=3),
                        gray_level=64,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'MP: Lenovo': Judger(
                name='MP: Lenovo',
                judgements=[
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=5),
                        gray_level=64,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=5),
                        gray_level=64,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'Tablet: Huawei': Judger(
                name='Tablet: Huawei',
                judgements=[
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=1),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'Tablet: Lenovo': Judger(
                name='Tablet: Lenovo',
                judgements=[
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=2),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=2),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'Tablet: 百度': Judger(
                name='Tablet: 百度',
                judgements=[
                    Judgement(
                        stress_time=timedelta(hours=2),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=2),
                        recover_time=timedelta(minutes=3),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'Tablet: 阿里': Judger(
                name='Tablet: 阿里',
                judgements=[
                    Judgement(
                        stress_time=timedelta(hours=6),
                        recover_time=timedelta(minutes=1),
                        gray_level=64,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=6),
                        recover_time=timedelta(minutes=1),
                        gray_level=64,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'Tablet: Google': Judger(
                name='Tablet: Google',
                judgements=[
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(seconds=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=5),
                        recover_time=timedelta(seconds=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=4),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=4),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'Tablet: 傳音 for LG': Judger(
                name='Tablet: 傳音 for LG',
                judgements=[
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(seconds=5),
                        gray_level=64,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(seconds=5),
                        gray_level=64,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(seconds=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(seconds=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'Tablet: Amazon 10x16': Judger(
                name='Tablet: Amazon',
                judgements=[
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(minutes=2),
                        gray_level=32,
                        is_type=ISType.S,
                        ra_level=0,
                    ),
                    Judgement(
                        stress_time=timedelta(minutes=30),
                        recover_time=timedelta(minutes=2),
                        gray_level=32,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            ),
            'MP: HP': Judger(
                name='MP: HP',
                judgements=[
                    Judgement(
                        stress_time=timedelta(hours=24),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.S,
                        ra_level=2,
                    ),
                    Judgement(
                        stress_time=timedelta(hours=24),
                        recover_time=timedelta(minutes=5),
                        gray_level=128,
                        is_type=ISType.L,
                        ra_level=0,
                    ),
                ]
            )
        }
        
        return dummy_judgers
    
    @classmethod
    def table(
        cls,
        ws: Worksheet,
        logs: Dict[str, Dict[str, Judger]],
    ):
        colors = {
            'title_bg': Color.WY.value,
            # 'header_bg': ,
            'font': Color.K.value,
            'bg': Color.W.value,
        }
        
        ws.title = 'Result'
        cell_ptr = CellPointer(row=1, col=1)
        
        for title, log in logs.items():
            
            # title
            cell_ptr.col = 1
            title_cell = ws.cell(*cell_ptr)
            title_cell.value = title
            title_cell.fill = PatternFill(
                "solid", fgColor=colors['title_bg']
            )
            title_cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
            )
            cell_ptr.row += 1
            
            # header
            ws.cell(*cell_ptr).value = 'Condition'
            ws.cell(*cell_ptr).alignment = Alignment(
                horizontal='center',
                vertical='center',
            )
            ws.row_dimensions[cell_ptr.row].height = 70
            cell_ptr.col += 1
            value = ''
            for test in cls.load_specs()[title].judgements:
                if test.is_type == ISType.S:
                    value += f'stress time: {test.stress_time}\n'
                    value += f'recover time: {test.recover_time}\n'
                    value += f'gray level: {test.gray_level}\n'
                    value += f'≤ S{test.ra_level}\n'
                    col_len = len(f'recover time: {test.recover_time}')
                if test.is_type == ISType.L:
                    value += f'≤ L{test.ra_level}'
                    ws.cell(*cell_ptr).value = value
                    ws.column_dimensions[
                        get_column_letter(cell_ptr.col)
                    ].width = col_len
                    
                    value = ''
                    cell_ptr.col += 1
            cell_ptr.row += 1
            # using the col number to merge the title cell
            ws.merge_cells(
                start_row=title_cell.row,
                start_column=title_cell.column,
                end_row=title_cell.row,
                end_column=cell_ptr.col - 1,
            )
            
            # value
            for condition, jugder in log.items():
                cell_ptr.col = 1
                ws.cell(*cell_ptr).value = condition
                cell_ptr.col += 1
                value = ''
                is_pass = True
                for jugement in jugder.judgements:
                    value += jugement.yeild + ' '
                    value += jugement.is_type.name
                    value += str(jugement.ng_level)
                    is_pass = jugement.is_pass and is_pass
                
                    if jugement.is_type == ISType.S:
                        value += ', '
                        
                    if jugement.is_type == ISType.L:
                        if '0/0' in value:
                            is_pass = None
                        if '0/0' in value:
                            is_pass = None
                        if is_pass is True:
                            colors['font'] =  Color.K.value
                            colors['bg'] = Color.G.value
                        if is_pass is False:
                            colors['bg'] = Color.R.value
                            colors['font'] = Color.W.value
                        if is_pass is None:
                            colors['font'] = Color.K.value
                            colors['bg'] = Color.Y.value
                            
                        ws.cell(*cell_ptr).value = value
                        ws.cell(*cell_ptr).fill = PatternFill(
                            "solid", fgColor=colors['bg']
                        )
                        ws.cell(*cell_ptr).font = Font(
                            name='Calibri', color=colors['font']
                        )
                        value = ''
                        is_pass = True
                        cell_ptr.col += 1
                    
                cell_ptr.row += 1            
            cell_ptr.row += 1
                    
class Parser:
    def __init__(self, workbook: Workbook):
        self.workbook = workbook
        self.logs = {}
    
    def parse(self) -> List[Log]:
        result = []
        for condition in self.workbook.sheetnames:
            result += self.parse_condition(condition)
        return result
    
    @staticmethod
    def generate_checkpoints(
        stress_time: timedelta,
        recover_times: List[timedelta],
        gray_levels: List[int],
    ) -> List[CheckPoint]:
        result: list[CheckPoint] = []
        for gray_level in gray_levels:
            for is_type in ISType:
                for recover_time in recover_times:
                    result.append(
                        CheckPoint(
                            stress_time=stress_time,
                            recover_time=recover_time,
                            gray_level=gray_level,
                            is_type=is_type,
                        )
                    )
                
        return result
    
    def parse_condition(
        self,
        condition: str,
    ) -> List[Log]:
        
        logs = {}
        for row in self.workbook[condition]['C9:D13']:
            if row[0].value is None:
                continue
            logs[row[0].value] = Log(
                name=row[0].value,
                remark=row[1].value,
                condition=condition,
                value=[],
            )
        
        for section in self.sections:
            self.parse_section(
                logs, condition,
                **section
            )
        
        return [log for _, log in logs.items()]
            
    def parse_section(
        self,
        logs: Dict[str, Log],
        condition: str,
        row: int,
        checkpoints: List[CheckPoint],
    ):        
        cell_range = f"C{row}:P{row+4}"

        for row in self.workbook[condition][cell_range]:
            if row[0].value is None:
                continue
            log = logs[row[0].value]
            
            index: int = 1
            seen: set[tuple[timedelta, int, ISType]] = set()
            for checkpoint in checkpoints:
                if (without_recover_time := (
                    checkpoint.stress_time,
                    checkpoint.gray_level,
                    checkpoint.is_type,
                )) not in seen:
                    index += 1
                    recover_index = 0
                
                seen.add(without_recover_time)
                try:
                    allowed_log = re.findall(r'\d', str(row[index].value))
                    if recover_index < len(allowed_log):
                        ra_level = int(allowed_log[recover_index])
                    else:
                        ra_level = int(allowed_log[-1])
                except:
                    continue
                
                recover_index += 1
                
                log.value.append(
                    ImageStickingBase(
                        **checkpoint.dict(),
                        ra_level=ra_level,
                    )
                )
            
                
    @property
    def sections(self) -> Dict:
        sections = [
            {
                'row': 9,
                'checkpoints': self.generate_checkpoints(
                    timedelta(seconds=10), 
                    [timedelta(seconds=0)],
                    [0, 39, 64, 128],
                ) + self.generate_checkpoints(
                    timedelta(minutes=3),
                    [timedelta(seconds=0), timedelta(minutes=1)],
                    [0, 39, 64, 128],
                ),
            },
            {
                'row': 18,
                'checkpoints': self.generate_checkpoints(
                    timedelta(minutes=5),
                    [
                        timedelta(seconds=0), 
                        timedelta(seconds=5), 
                        timedelta(seconds=10)
                    ],
                    [0, 39, 32, 64, 128],
                ) + self.generate_checkpoints(
                    timedelta(minutes=10),
                    [timedelta(seconds=0), timedelta(seconds=5)],
                    [0, 128],
                )
            },
            {
                'row': 27,
                'checkpoints': self.generate_checkpoints(
                    timedelta(minutes=30),
                    [
                        timedelta(seconds=0), 
                        timedelta(minutes=2),
                        timedelta(minutes=3),
                    ],
                    [32, 64, 128],
                )
            },
            {
                'row': 36,
                'checkpoints': self.generate_checkpoints(
                    timedelta(hours=1),
                    [
                        timedelta(seconds=0),
                        timedelta(seconds=2),
                        timedelta(seconds=20), 
                        timedelta(minutes=3), 
                        timedelta(minutes=5),
                        timedelta(minutes=10),
                    ],
                    [32, 51, 64, 128],
                )
            },
            {
                'row': 45,
                'checkpoints': self.generate_checkpoints(
                    timedelta(hours=2),
                    [
                        timedelta(seconds=0),
                        timedelta(minutes=3), 
                        timedelta(minutes=5),
                    ],
                    [128],
                ) + self.generate_checkpoints(
                    timedelta(hours=4),
                    [
                        timedelta(seconds=0), 
                        timedelta(minutes=5)
                    ],
                    [0, 39, 128],
                )
            },
            {
                'row': 54,
                'checkpoints': self.generate_checkpoints(
                    timedelta(hours=6),
                    [
                        timedelta(seconds=0), 
                        timedelta(minutes=1), 
                        timedelta(minutes=10),
                    ],
                    [64, 128],
                )
            },
            {
                'row': 63,
                'checkpoints': self.generate_checkpoints(
                    timedelta(hours=24),
                    [
                        timedelta(seconds=0), 
                        timedelta(minutes=5),
                    ],
                    [0, 39, 64, 128],
                )
            },
        ]
        
            
        return sections
    
    def load_setting(self, path: Path):
        ...