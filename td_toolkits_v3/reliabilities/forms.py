from django import forms
from django.http import HttpRequest
from django.core.cache import cache

import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime

from td_toolkits_v3.materials.models import (
    LiquidCrystal,
    Polyimide,
    Seal,
    Vender,
)

from .models import (
    Batch,
    Adhesion,
    DeltaAngle,
    File,
    LowTemperatureOperation,
    LowTemperatureStorage,
    PressureCookingTest,
    SealWVTR,
    UShapeAC,
    VoltageHoldingRatio,
)

from .tools import utils
from .tools import image_sticking

class ReliabilitiesUploadForm(forms.Form):
    reliabilites = forms.FileField(
        help_text='Excel file',
        widget=forms.FileInput(attrs={'accept': '.xlsx'})
    )

    @staticmethod
    def get_configureation(row):
        """
        Simplify the way to get the whole configuration.
        Including deal with create and None value
        
        Parameters
        ----------
        row: list-like array
            The structure should ('', lc, pi, seal, ...)
        """
        if row[1] == 'N.A.':
            lc = None
        else:
            lc, _ = LiquidCrystal.objects.get_or_create(name=row[1])
        if row[2] == 'N.A.':
            pi = None
        else:
            pi, _ = Polyimide.objects.get_or_create(name=row[2])
        if row[3] == 'N.A.':
            seal = None
        else:
            seal, _ = Seal.objects.get_or_create(name=row[3])
        return lc, pi, seal

    def save(self):
        # TODO: This should also have a more elegant way to implment it.
        adhesion_df = pd.read_excel(
            self.cleaned_data['reliabilites'], sheet_name='Adhesion'
        )
        lto_df = pd.read_excel(
            self.cleaned_data['reliabilites'], sheet_name='LTO'
        )
        lts_df = pd.read_excel(
            self.cleaned_data['reliabilites'], sheet_name='LTS'
        )
        pct_df = pd.read_excel(
            self.cleaned_data['reliabilites'], sheet_name='PCT'
        )
        seal_wvtr_df = pd.read_excel(
            self.cleaned_data['reliabilites'], sheet_name='Seal WVTR'
        )
        delta_angle_df = pd.read_excel(
            self.cleaned_data['reliabilites'], sheet_name='Δangle'
        )
        u_shape_ac_df = pd.read_excel(
            self.cleaned_data['reliabilites'], sheet_name='U-Shape AC'
        )
        vhr_df = pd.read_excel(
            self.cleaned_data['reliabilites'], sheet_name='VHR'
        )
        # Store each table
        if len(adhesion_df):
            # TODO: The field would change later
            for _, row in adhesion_df.iterrows():
                lc, pi, seal = self.get_configureation(row)
                vender, _ = Vender.objects.get_or_create(name=row[9])
                file, _ = File.objects.get_or_create(name=row[10])
                batch, _ = Batch.objects.get_or_create(name=row[11])
                if not Adhesion.objects.filter(
                    lc=lc, pi=pi, seal=seal, vender=vender, file_source=file,
                    adhesion_interface=row[6], method=row[7], batch=batch).exists():
                    Adhesion.objects.create(
                        lc=lc, pi=pi, seal=seal, vender=vender, 
                        file_source=file,
                        unit=row[5],
                        adhesion_interface=row[6], 
                        method=row[7], 
                        value=row[4], 
                        peeling=row[8],
                        batch=batch
                    )

        if len(lto_df):
            for _, row in lto_df.iterrows():
                lc, pi, seal = self.get_configureation(row)
                vender, _ = Vender.objects.get_or_create(name=row[9])
                file, _ = File.objects.get_or_create(name=row[10])
                jar_test_seal, _ = Seal.objects.get_or_create(name=row[7])
                batch, _ = Batch.objects.get_or_create(name=row[11])
                
                if not LowTemperatureOperation.objects.filter(
                    lc=lc, pi=pi, seal=seal, vender=vender, file_source=file,
                    storage_condition=row[5], slv_condition=row[6],
                    jar_test_seal=jar_test_seal,
                    measure_temperature=row[8], batch=batch).exists():

                    LowTemperatureOperation.objects.create(
                        lc=lc, pi=pi, seal=seal, vender=vender, file_source=file,
                        storage_condition=row[5], slv_condition=row[6],
                        jar_test_seal=jar_test_seal, measure_temperature=row[8],
                        value=LowTemperatureOperation.value_mapping[row[4]],
                        batch=batch
                    )

        if len(lts_df):
            for _, row in lts_df.iterrows():
                lc, pi, seal = self.get_configureation(row)
                vender, _ = Vender.objects.get_or_create(name=row[9])
                file, _ = File.objects.get_or_create(name=row[10])
                batch, _ = Batch.objects.get_or_create(name=row[11])
                if (row[7] != 'N.A.'):
                    jar_test_seal, _ = Seal.objects.get_or_create(name=row[7])
                else:
                    jar_test_seal = None
                
                if not LowTemperatureStorage.objects.filter(
                    lc=lc, pi=pi, seal=seal, vender=vender, file_source=file,
                    storage_condition=row[5], slv_condition=row[6],
                    jar_test_seal=jar_test_seal,
                    measure_temperature=row[8], batch=batch).exists():
                    LowTemperatureStorage.objects.create(
                        lc=lc, pi=pi, seal=seal, vender=vender, 
                        file_source=file,
                        storage_condition=row[5], slv_condition=row[6],
                        jar_test_seal=jar_test_seal, measure_temperature=row[8],
                        value=row[4], batch=batch)

        if len(pct_df):
            for _, row in pct_df.iterrows():
                lc, pi, seal = self.get_configureation(row)
                vender, _ = Vender.objects.get_or_create(name=row[7])
                file, _ = File.objects.get_or_create(name=row[8])
                batch, _ = Batch.objects.get_or_create(name=row[9])
                if not PressureCookingTest.objects.filter(
                    lc=lc, pi=pi, seal=seal, vender=vender, file_source=file,
                    measure_condition=row[5], test_vehical=row[6],
                    batch=batch,
                    ).exists():
                    PressureCookingTest.objects.create(
                        value=row[4], lc=lc, pi=pi, seal=seal, vender=vender, 
                        file_source=file,
                        measure_condition=row[5], test_vehical=row[6],
                        batch=batch,
                    )

        if len(seal_wvtr_df):
            # TODO: The field would change
            for _, row in seal_wvtr_df.iterrows():
                lc, pi, seal = self.get_configureation(row)
                vender, _ = Vender.objects.get_or_create(name=row[10])
                file, _ = File.objects.get_or_create(name=row[11])
                batch, _ = Batch.objects.get_or_create(name=row[12])
                if not SealWVTR.objects.filter(
                    lc=lc, pi=pi, seal=seal, vender=vender, file_source=file,
                    temperature=row[7], humidity=row[8], thickness=row[9],
                    batch=batch
                ).exists():
                    SealWVTR.objects.create(
                        value=row[4], lc=lc, pi=pi, seal=seal, vender=vender, 
                        file_source=file,
                        unit=row[5], time=row[6],
                        temperature=row[7], humidity=row[8], thickness=row[9],
                        batch=batch
                    )

        if len(delta_angle_df):
            for _, row in delta_angle_df.iterrows():
                lc, pi, seal = self.get_configureation(row)
                vender, _ = Vender.objects.get_or_create(name=row[9])
                file, _ = File.objects.get_or_create(name=row[10])
                batch, _ = Batch.objects.get_or_create(name=row[11])
                if not DeltaAngle.objects.filter(
                    lc=lc, pi=pi, seal=seal, vender=vender, file_source=file,
                    measure_voltage=row[5], measure_freq=row[6], 
                    measure_time=row[7],
                    measure_temperature=row[8],
                    batch=batch
                ).exists():
                    DeltaAngle.objects.create(
                        value=row[4], lc=lc, pi=pi, seal=seal, vender=vender, 
                        file_source=file,
                        measure_voltage=row[5], 
                        measure_freq=row[6], 
                        measure_time=row[7],
                        measure_temperature=row[8],
                        batch=batch
                    )

        if len(u_shape_ac_df):
            for _, row in u_shape_ac_df.iterrows():
                lc, pi, seal = self.get_configureation(row)
                vender, _ = Vender.objects.get_or_create(name=row[7])
                file, _ = File.objects.get_or_create(name=row[8])
                batch, _ = Batch.objects.get_or_create(name=row[9])
                if not UShapeAC.objects.filter(
                    lc=lc, pi=pi, seal=seal, vender=vender, file_source=file,
                    time=row[5], temperature=row[6],batch=batch).exists():
                    UShapeAC.objects.create(
                        value=row[4], lc=lc, pi=pi, seal=seal, vender=vender, 
                        file_source=file,
                        time=row[5], temperature=row[6],batch=batch)

        if len(vhr_df):
            for _, row in vhr_df.iterrows():
                lc, pi, seal = self.get_configureation(row)
                vender, _ = Vender.objects.get_or_create(name=row[9])
                file, _ = File.objects.get_or_create(name=row[10])
                batch, _ = Batch.objects.get_or_create(name=row[11])
                if not VoltageHoldingRatio.objects.filter(
                    lc=lc, pi=pi, seal=seal, vender=vender, file_source=file,
                    measure_voltage=row[5], measure_freq=row[6], 
                    measure_temperature=row[7],
                    uv_aging=row[8], batch=batch).exists():
                    VoltageHoldingRatio.objects.create(
                        value=row[4], lc=lc, pi=pi, seal=seal, vender=vender, 
                        file_source=file,
                        measure_voltage=row[5], measure_freq=row[6], 
                        measure_temperature=row[7],
                        uv_aging=row[8], batch=batch
                    )

class ReliabilityPhaseTwoForm(forms.Form):
    batch = forms.ModelChoiceField(queryset=Batch.objects.all())
    vhr_weight = forms.FloatField(label='VHR(%) weight', initial=1.)
    delta_angle_weight = forms.FloatField(label='Δ angle(°) weight', initial=1.)
    adhesion_weight = forms.FloatField(label='Adhesion(kgw) weight', initial=1.)
    lts_weight = forms.FloatField(label='LTS(days) weight', initial=1.)
    pct_weight = forms.FloatField(label='PCT(hrs) weight', initial=1.)
    sealwvtr_weight = forms.FloatField(label='Seal WVTR(%) weight', initial=1.)
    ushape_ac_weight = forms.FloatField(label='U-Shape AC(%) weight', initial=1.)
    
    def calc(self, request: HttpRequest):
        weight = {
            'VHR(heat)': self.cleaned_data['vhr_weight'],
            'Δ angle': self.cleaned_data['delta_angle_weight'],
            'Adhesion test': self.cleaned_data['adhesion_weight'],
            'LTS': self.cleaned_data['lts_weight'],
            'PCT': self.cleaned_data['pct_weight'],
            'Seal WVTR': self.cleaned_data['sealwvtr_weight'],
            'U-shape AC%': self.cleaned_data['ushape_ac_weight'],
        }
        batch = self.cleaned_data['batch']
        
        result = utils.ReliabilityPhaseTwo(batch)
        result.scale_factor = weight
        
        result.search(VoltageHoldingRatio)
        result.search(DeltaAngle)
        result.search(Adhesion)
        result.search(LowTemperatureStorage)
        result.search(PressureCookingTest)
        result.search(SealWVTR)
        result.search(UShapeAC)
        
        request.session['summary'] = result.summary.to_json()
        request.session['score'] = result.score.to_json()
        request.session['plot'] = result.plot
    
class ImageStickingUploadForm(forms.Form):
    file = forms.FileField(
        help_text='Excel file(.xlsx)',
        widget=forms.FileInput(attrs={'accept': '.xlsx'})
    )
    
    def calc(self):
        wb = load_workbook(self.cleaned_data['file'])
        file_name = (
            # f"{self.cleaned_data['file'].name}_"
            "traffic_light_"
            f"{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        )
        
        parser = image_sticking.Parser(wb)
        result = parser.parse()
        conditions = pd.DataFrame(
            [i.dict(exclude={'value'}) for i in result]
        )['condition'].unique()
        
        logs = {}
        for judge_kind, judger in image_sticking.Judger.load_specs().items():
            logs[judge_kind] = {}
            for condition in conditions:
                log = [i for i in filter(lambda x: x.condition==condition, result)]
                logs[judge_kind][condition] = image_sticking.Judger(
                    name=condition,
                    judgements=judger.judgements,
                )
                logs[judge_kind][condition].judge(log)
                    
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
                    
        image_sticking.Judger.table(ws, logs)
        cache.set('file_name', file_name)
        cache.set('wb', wb)